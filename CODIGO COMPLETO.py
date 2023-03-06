while True:
    import imaplib
    import email
    import openpyxl
    import os
    
    try:
        mail = imaplib.IMAP4_SSL('imap.gmail.com')
        #mail.login('lucasbenjamin832@gmail.com', 'vqmbokydlbzggodj')
        #mail.login('luc45olivare5@gmail.com', 'vnogcftoyqxrffyk')--
        mail.login('paolachilquinta@gmail.com', 'qgcxxagoyqknumxn') #contraseña gmail: chilquinta
        mail.select('inbox')
    except:
        print("Quizas no tiene internet, intente de nuevo")   
        break 

    subject = input("¿Cual es el asunto del correo?: ")
    print("")
    print("####################################################################################################################")
    print("Muy bien, espere un momento :)")
    print("####################################################################################################################")
    print("")
    
    print("####################################################################################################################")
    print("Buscando mensaje") 
    print("##############################################p######################################################################")
    print("")
    status, messages = mail.search(None, f'SUBJECT "{subject}"') 
    messages = messages[0].split()
    if messages:
        print("####################################################################################################################")
        print("El asunto del correo ha sido encontrado")
        print("####################################################################################################################")
        print("")
    else:
        print("####################################################################################################################")
        print("\033[4;31m" + "El asunto del correo no ha sido encontrado" + "\033[0;m")
        print("####################################################################################################################")
        print("")
    for message in messages:
        status, msg = mail.fetch(message, '(RFC822)')
        for response in msg:
            if isinstance(response, tuple):
                msg = email.message_from_bytes(response[1])
                body = None
                for part in msg.walk():
                    if part.get_content_type() == 'text/plain':
                        body = part.get_payload(decode=True).decode()
                        break
                if body:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    data = body.splitlines()
                    row = 1
                    col = 1
                    for item in data:
                        sheet.cell(row=row, column=col).value = item
                        col += 1
                        if col > 2:
                            col = 1 
                            row += 1
                    wb.save("TABLA ELECTRODEPENDIENTES.xlsx")
                    
                    
                    if os.path.exists("TABLA ELECTRODEPENDIENTES STAR.xlsx"):
                        wb2 = openpyxl.load_workbook("TABLA ELECTRODEPENDIENTES STAR.xlsx")
                        sheet2 = wb2.active
                        last_row = sheet2.max_row
                        print("####################################################################################################################")
                        print("Llevando datos a TABLA ELECTRODEPENDIENTES STAR")     
                        print("####################################################################################################################")
                    else:
                        print("\033[4;31m" + "No se ha encontrado TABLA ELECTRODEPENDIENTES STAR" + "\033[0;m")
                        if os.path.exists("TABLA ELECTRODEPENDIENTES.xlsx"):
                            os.remove("TABLA ELECTRODEPENDIENTES.xlsx")     
                        print("")
                        print("####################################################################################################################")
                        print("")
                        break
                    wb2.save("TABLA ELECTRODEPENDIENTES STAR.xlsx")

                    for i in range(1, 20):  
                        sheet2.cell(row=3541, column=i + 3).value = sheet.cell(row=i, column=2).value

                        #DATOS TABLA NO ACTUALIZADA
    
                        sheet2.cell(row=last_row + 1, column=6).value = sheet2.cell(row=3541, column=20).value
                        sheet2.cell(row=last_row + 1, column=10).value = sheet2.cell(row=3541, column=5).value
                        sheet2.cell(row=last_row + 1, column=21).value = sheet2.cell(row=3541,  column=18).value
                        sheet2.cell(row=last_row + 1, column=19).value = sheet2.cell(row=3541, column=15).value
                        sheet2.cell(row=last_row + 1, column=20).value = sheet2.cell(row=3541, column=17).value
                        sheet2.cell(row=last_row + 1, column=4).value = str(sheet2.cell(row=3541, column=7).value) + "/" + str(sheet2.cell(row=3541, column=8).value)
                        sheet2.cell(row=last_row + 1, column=2).value = "122022" #NO MODIFICAR
                        sheet2.cell(row=last_row + 1, column=3).value = "****"#MANUAL
                        sheet2.cell(row=last_row + 1, column=5).value = "S/N" #NO MODIFICAR
                        sheet2.cell(row=last_row + 1, column=6).value = sheet2.cell(row=3541, column=9).value
                        sheet2.cell(row=last_row + 1, column=8).value = "1" #NO MODIFICAR
                        sheet2.cell(row=last_row + 1, column=9).value = "1" #NO MODIFICAR
                        sheet2.cell(row=last_row + 1, column=11).value = "****"#MANUAL
                        sheet2.cell(row=last_row + 1, column=1).value = "006"#MANUAL

                        word = sheet2.cell(row=3541, column=4).value.split()[1]
                        sheet2.cell(row=last_row + 1, column=12).value = word

                        word = sheet2.cell(row=3541, column=4).value.split()[2]
                        sheet2.cell(row=last_row + 1, column=13).value = word

                        word = sheet2.cell(row=3541, column=4).value.split()[0]
                        sheet2.cell(row=last_row + 1, column=14).value = word

                        sheet2.cell(row=last_row + 1, column=15).value = "FECHA 1" #(Manual)
                        sheet2.cell(row=last_row + 1, column=16).value = "FECHA 2" #(Manual)
                        sheet2.cell(row=last_row + 1, column=17).value = "1 NUMERO" #(Manual)
                        sheet2.cell(row=last_row + 1, column=18).value = "1 NUMERO" #(Manual)

                        #DATOS TABLA ACTUALIZADA 
                        sheet2.cell(row=3543, column=6).value = sheet2.cell(row=3541, column=4).value #NOMBRE DEL PACIENTE #F
                        sheet2.cell(row=3543, column=7).value = sheet2.cell(row=3541, column=5).value #RUT PE SIN PUNTOS Y CON GUION #G
                        sheet2.cell(row=3543, column=10).value = sheet2.cell(row=3541, column=7).value #DIRECCIÓN #J
                        sheet2.cell(row=3543, column=5).value = sheet2.cell(row=3541, column=8).value #COMUNA #E
                        sheet2.cell(row=3543, column=2).value = sheet2.cell(row=3541, column=9).value #NUMERO DE PRODUCTO #B 
                        sheet2.cell(row=3543, column=12).value = sheet2.cell(row=3541, column=15).value #NOMBRE Y APELLIDOS TUTOR #L                 
                        sheet2.cell(row=3543, column=8).value = str(sheet2.cell(row=3541, column=6).value) + "/" + str(sheet2.cell(row=last_row + 1, column=20).value) #Teléfono PE #H Y #TELEFONO TUTOR #H
                        sheet2.cell(row=3543, column=9).value = str(sheet2.cell(row=3541, column=11).value) + "/" + str(sheet2.cell(row=last_row + 1, column=21).value) #CORREO ELÉCTRONICO PE #I Y #CORREO ELÉCTRONICO TUTOR I

                    wb2.save("TABLA ELECTRODEPENDIENTES STAR.xlsx")
                    sheet2.delete_rows(3541, 1)
                    wb2.save("TABLA ELECTRODEPENDIENTES STAR.xlsx")

                    if os.path.exists("TABLA ELECTRODEPENDIENTES.xlsx"):
                        os.remove("TABLA ELECTRODEPENDIENTES.xlsx")
                    wb2.save("TABLA ELECTRODEPENDIENTES STAR.xlsx")
                    print("")
                    print("####################################################################################################################")
                    print("Los datos han sido copiados en TABLA ELECTRODEPENDIENTES STAR") 
                    print("####################################################################################################################")
                    
                  # Abrir los archivos de Excel 
                    archivo1 = openpyxl.load_workbook("TABLA ELECTRODEPENDIENTES STAR.xlsx")

                    
                    if os.path.exists("ED ACTUALIZADA CONSOLIDADA.xlsx"):
                        archivo2 = openpyxl.load_workbook("ED ACTUALIZADA CONSOLIDADA.xlsx")
                        print("")
                        print("####################################################################################################################")
                        print("Abriendo tabla ED ACTUALIZADA CONSOLIDADA") 
                        print("####################################################################################################################")
                    else:
                        print("")
                        print("####################################################################################################################")
                        print("")
                        print("\033[4;31m" + "No se ha encontrado la tabla ED ACTUALIZADA CONSOLIDADA" + "\033[0;m")
                        print("")
                        print("####################################################################################################################")
                        print("")
                        break

                    # Seleccionar la hoja de cálculo de cada archivo
                    hoja1 = archivo1.active
                    hoja2 = archivo2.active

                    # Definir la fila a copiar
                    fila_a_copiar = 3542

                    # Copiar los valores de cada celda de la fila especificada
                    fila_copiada = [hoja1.cell(row=fila_a_copiar, column=columna).value for columna in range(1, hoja1.max_column + 1)]

                    
                    # Obtener la última fila de datos en la hoja2
                    ultima_fila = hoja2.max_row + 1

                    for columna, valor in enumerate(fila_copiada, 1):
                        hoja2.cell(row=ultima_fila, column=columna, value=valor)

                        # Carga el archivo de Excel existente "TABLA ELECTRODEPENDIENTES_11202.xlsx"

                    wb = openpyxl.load_workbook("ED ACTUALIZADA CONSOLIDADA.xlsx")      
                    hoja1 = wb.active
                    print("")
                    print("####################################################################################################################")
                    print("Los datos se han llevado con exito")
                    print("####################################################################################################################")
                
                    # Save changes to workbook
                    wb.save("ED ACTUALIZADA CONSOLIDADA.xlsx") 

                    # Guardar los cambios en el archivo 2
                    archivo2.save("ED ACTUALIZADA CONSOLIDADA.xlsx")

                    wb2 = openpyxl.load_workbook("TABLA ELECTRODEPENDIENTES STAR.xlsx")
                    sheet2 = wb2.active  

                    sheet2.delete_rows(3542, 1)
                    wb2.save("TABLA ELECTRODEPENDIENTES STAR.xlsx")
                    print("")
                    print("####################################################################################################################")
                    print("🌟 TABLAS COPIADAS CON ÉXITO 🌟")    
                    print("####################################################################################################################")
                    print("")
                    print("")
                    print("")

    mail.close()
    mail.logout()

#día 15 de febrero, estuvo buena 21 cromosomas le doy un 9.5/10 no le doy el 10 porque creo que hace falta una temporada 2
#dia 16 de febrero, esta interesante narcos, es una buena serie
#día 17 de febrero, narcos primera temporada me parece muy buena ya me terminé la 1era temporada, voy por la segunda 
#día 20 de febrero, narcos segunda temporada meh, no esta mal, pero no esta mal. Me puse a estudiar otra cosa
#día 21 de febrero, llevo 2 semanas sin hacer nada, odio a los clientes, ojala les caiga un rayo 
